"""
FSAE Data Post-Processing Script
Generates Excel workbooks with graphs from CSV sensor data
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from pathlib import Path
import tkinter as tk
from tkinter import ttk

def select_columns_dialog(columns, title_text="Select which columns to create graphs for:"):
    """
    Create a dialog for user to select which columns to graph
    
    Args:
        columns: List of column names
        title_text: Title text for the dialog
    
    Returns:
        List of selected column names
    """
    selected_columns = []
    
    def on_submit():
        nonlocal selected_columns
        selected_columns = [col for col, var in checkboxes.items() if var.get()]
        root.quit()
        root.destroy()
    
    def select_all():
        for var in checkboxes.values():
            var.set(True)
    
    def deselect_all():
        for var in checkboxes.values():
            var.set(False)
    
    # Create main window
    root = tk.Tk()
    root.title("Select Columns")
    root.geometry("600x700")
    
    # Title label
    title_label = tk.Label(root, text=title_text, 
                           font=("Arial", 12, "bold"))
    title_label.pack(pady=10)
    
    # Button frame
    button_frame = tk.Frame(root)
    button_frame.pack(pady=5)
    
    select_all_btn = tk.Button(button_frame, text="Select All", command=select_all)
    select_all_btn.pack(side=tk.LEFT, padx=5)
    
    deselect_all_btn = tk.Button(button_frame, text="Deselect All", command=deselect_all)
    deselect_all_btn.pack(side=tk.LEFT, padx=5)
    
    # Create scrollable frame for checkboxes
    canvas = tk.Canvas(root)
    scrollbar = ttk.Scrollbar(root, orient="vertical", command=canvas.yview)
    scrollable_frame = ttk.Frame(canvas)
    
    scrollable_frame.bind(
        "<Configure>",
        lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
    )
    
    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)
    
    # Create checkboxes
    checkboxes = {}
    for col in columns:
        if col != 'Time (sec)':  # Don't include time column
            var = tk.BooleanVar(value=True)  # Default to selected
            checkboxes[col] = var
            cb = ttk.Checkbutton(scrollable_frame, text=col, variable=var)
            cb.pack(anchor='w', padx=20, pady=2)
    
    canvas.pack(side="left", fill="both", expand=True, padx=(20, 0))
    scrollbar.pack(side="right", fill="y")
    
    # Submit button
    submit_btn = tk.Button(root, text="Continue", command=on_submit, 
                          bg="green", fg="white", font=("Arial", 10, "bold"))
    submit_btn.pack(pady=20)
    
    root.mainloop()
    
    return selected_columns

def get_scaling_factors_dialog(columns):
    """
    Create a dialog for user to input scaling factors for selected columns
    
    Args:
        columns: List of column names
    
    Returns:
        Dictionary mapping column names to scaling factors
    """
    scaling_factors = {}
    
    def on_submit():
        nonlocal scaling_factors
        for col, entry in entries.items():
            try:
                value = float(entry.get())
                scaling_factors[col] = value
            except ValueError:
                scaling_factors[col] = 1.0  # Default to 1.0 if invalid
        root.quit()
        root.destroy()
    
    def set_all_to_one():
        for entry in entries.values():
            entry.delete(0, tk.END)
            entry.insert(0, "1.0")
    
    # Create main window
    root = tk.Tk()
    root.title("Set Scaling Factors")
    root.geometry("600x700")
    
    # Title label
    title_label = tk.Label(root, text="Enter scaling factors for each column:", 
                           font=("Arial", 12, "bold"))
    title_label.pack(pady=10)
    
    info_label = tk.Label(root, text="(Values will be multiplied by the scaling factor)", 
                          font=("Arial", 9))
    info_label.pack(pady=5)
    
    # Button to set all to 1
    set_one_btn = tk.Button(root, text="Set All to 1.0", command=set_all_to_one)
    set_one_btn.pack(pady=5)
    
    # Create scrollable frame for entries
    canvas = tk.Canvas(root)
    scrollbar = ttk.Scrollbar(root, orient="vertical", command=canvas.yview)
    scrollable_frame = ttk.Frame(canvas)
    
    scrollable_frame.bind(
        "<Configure>",
        lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
    )
    
    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)
    
    # Create entry fields for each column
    entries = {}
    for col in columns:
        frame = ttk.Frame(scrollable_frame)
        frame.pack(fill='x', padx=20, pady=5)
        
        label = ttk.Label(frame, text=col, width=40)
        label.pack(side='left')
        
        entry = ttk.Entry(frame, width=15)
        entry.insert(0, "1.0")  # Default scaling factor
        entry.pack(side='right', padx=10)
        
        entries[col] = entry
    
    canvas.pack(side="left", fill="both", expand=True, padx=(20, 0))
    scrollbar.pack(side="right", fill="y")
    
    # Submit button
    submit_btn = tk.Button(root, text="Continue", command=on_submit, 
                          bg="green", fg="white", font=("Arial", 10, "bold"))
    submit_btn.pack(pady=20)
    
    root.mainloop()
    
    return scaling_factors

def select_x_axis_dialog():
    """
    Create a dialog for user to select X-axis (Time, RPM, or Both)
    
    Returns:
        List of strings: ['Time (sec)'], ['RPM'], or ['Time (sec)', 'RPM']
    """
    selected_axes = ["Time (sec)"]
    
    def on_time():
        nonlocal selected_axes
        selected_axes = ["Time (sec)"]
        root.quit()
        root.destroy()
    
    def on_rpm():
        nonlocal selected_axes
        selected_axes = ["RPM"]
        root.quit()
        root.destroy()
    
    def on_both():
        nonlocal selected_axes
        selected_axes = ["Time (sec)", "RPM"]
        root.quit()
        root.destroy()
    
    root = tk.Tk()
    root.title("Select X-Axis")
    root.geometry("500x250")
    
    label = tk.Label(root, text="Select X-axis for graphs:", 
                     font=("Arial", 12, "bold"))
    label.pack(pady=15)
    
    info_label = tk.Label(root, text="Choose whether to plot data against Time, RPM, or Both", 
                          font=("Arial", 9))
    info_label.pack(pady=5)
    
    button_frame = tk.Frame(root)
    button_frame.pack(pady=20)
    
    time_btn = tk.Button(button_frame, text="Time Only", command=on_time, 
                        bg="blue", fg="white", font=("Arial", 10, "bold"), 
                        width=12, height=2)
    time_btn.pack(side=tk.LEFT, padx=5)
    
    rpm_btn = tk.Button(button_frame, text="RPM Only", command=on_rpm, 
                       bg="green", fg="white", font=("Arial", 10, "bold"), 
                       width=12, height=2)
    rpm_btn.pack(side=tk.LEFT, padx=5)
    
    both_btn = tk.Button(button_frame, text="Both", command=on_both, 
                        bg="purple", fg="white", font=("Arial", 10, "bold"), 
                        width=12, height=2)
    both_btn.pack(side=tk.LEFT, padx=5)
    
    root.mainloop()
    
    return selected_axes

def get_combined_workbook_name():
    """
    Simple dialog to get the name for combined workbook
    
    Returns:
        String with workbook name
    """
    workbook_name = ""
    
    def on_submit():
        nonlocal workbook_name
        workbook_name = name_entry.get().strip()
        root.quit()
        root.destroy()
    
    root = tk.Tk()
    root.title("Combined Workbook Name")
    root.geometry("400x150")
    
    label = tk.Label(root, text="Enter name for combined workbook:", 
                     font=("Arial", 11))
    label.pack(pady=20)
    
    name_entry = tk.Entry(root, width=30, font=("Arial", 10))
    name_entry.insert(0, "Combined_Analysis")
    name_entry.pack(pady=10)
    
    submit_btn = tk.Button(root, text="OK", command=on_submit, 
                          bg="green", fg="white", font=("Arial", 10, "bold"))
    submit_btn.pack(pady=10)
    
    root.mainloop()
    
    return workbook_name if workbook_name else "Combined_Analysis"

def add_chart_to_sheet(wb, sheet_name, data_sheet, x_col, data_col, title, df, x_col_name, data_col_name):
    """
    Add a line chart to a new sheet in the workbook with data
    
    Args:
        wb: Workbook object
        sheet_name: Name for the new chart sheet
        data_sheet: Sheet containing the data
        x_col: Column index for x-axis (Time or RPM)
        data_col: Column index for data (y-axis)
        title: Chart title
        df: DataFrame containing the data
        x_col_name: Name of x-axis column (Time (sec) or RPM)
        data_col_name: Name of data column
    """
    # Create new sheet for chart
    chart_sheet = wb.create_sheet(sheet_name)
    
    # Copy relevant data to this sheet (starting at row 1)
    chart_sheet.cell(1, 1, x_col_name)
    chart_sheet.cell(1, 2, data_col_name)
    
    for idx, (x_val, data_val) in enumerate(zip(df[x_col_name], df[data_col_name]), start=2):
        chart_sheet.cell(idx, 1, x_val)
        chart_sheet.cell(idx, 2, data_val)
    
    # Create simple line chart (no fancy colors)
    chart = LineChart()
    chart.title = title
    chart.style = None  # Remove default styling
    chart.x_axis.title = x_col_name  # Use actual x-axis name (Time or RPM)
    chart.y_axis.title = title
    chart.width = 25   # Very large chart to give more room for axis titles
    chart.height = 15  # Very tall chart to give more room for axis titles
    
    # Show only ~5-6 labels across the entire x-axis (similar to y-axis)
    chart.x_axis.tickLblSkip = max(1, len(df) // 6)  # Show roughly 6 labels total
    chart.x_axis.number_format = '0'  # Format as whole numbers (no decimals)
    chart.x_axis.majorTickMark = "in"  # Add inside tick marks
    chart.x_axis.tickLblPos = "low"  # Position labels below axis
    
    # Add only horizontal gridlines (Y-axis gridlines only)
    chart.y_axis.majorGridlines = ChartLines()
    
    # Ensure axes are always visible
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    
    # Remove legend
    chart.legend = None
    
    # Add data series from the sheet itself
    y_values = Reference(chart_sheet, min_col=2, min_row=1, max_row=len(df)+1)
    x_values = Reference(chart_sheet, min_col=1, min_row=2, max_row=len(df)+1)
    
    chart.add_data(y_values, titles_from_data=True)
    chart.set_categories(x_values)
    
    # Simple line style - no colors
    if chart.series:
        chart.series[0].graphicalProperties.line.solidFill = "000000"  # Black line
        chart.series[0].smooth = True  # Smooth line
    
    # Add chart to sheet (positioned to the right of data)
    chart_sheet.add_chart(chart, "D2")

def process_csv_file(csv_path, selected_columns=None, scaling_factors=None, x_axes=None):
    """
    Process a single CSV file and generate Excel workbook with graphs
    
    Args:
        csv_path: Path to CSV file
        selected_columns: List of column names to graph (if None, will show dialog)
        scaling_factors: Dictionary mapping column names to scaling factors
        x_axes: List of X-axis column names (e.g., ['Time (sec)'], ['RPM'], or ['Time (sec)', 'RPM'])
    """
    if x_axes is None:
        x_axes = ['Time (sec)']
    
    print(f"\nProcessing: {csv_path}")
    
    # Read file and find where actual CSV data starts
    with open(csv_path, 'r', encoding='latin-1', errors='ignore') as f:
        lines = f.readlines()
    
    # Find the line that starts with "Time" (the header)
    header_line = 0
    for i, line in enumerate(lines):
        if line.startswith('Time'):
            header_line = i
            break
    
    # Read CSV starting from header line
    df = pd.read_csv(csv_path, encoding='latin-1', skiprows=header_line, on_bad_lines='skip')
    
    # Clean column names (remove extra spaces and invalid characters)
    df.columns = df.columns.str.strip()
    
    # Remove any columns with special characters or metadata
    valid_columns = []
    for col in df.columns:
        # Keep only columns that are proper sensor names (ASCII printable)
        if all(ord(c) < 128 for c in str(col)):
            valid_columns.append(col)
    
    df = df[valid_columns]
    
    # Remove rows with invalid data (metadata that got through)
    # Keep only rows where Time column is numeric
    df = df[pd.to_numeric(df['Time (sec)'], errors='coerce').notna()]
    
    # Reset index after filtering
    df = df.reset_index(drop=True)
    
    # Convert all columns to numeric where possible
    for col in df.columns:
        df[col] = pd.to_numeric(df[col], errors='ignore')
    
    # Show column selection dialog if not provided
    if selected_columns is None:
        print("  Opening column selection dialog...")
        selected_columns = select_columns_dialog(list(df.columns))
        
        if not selected_columns:
            print("  No columns selected. Skipping this file.\n")
            return
    
    # Apply scaling factors if provided
    if scaling_factors:
        for col in selected_columns:
            if col in df.columns and col in scaling_factors:
                scale = scaling_factors[col]
                if scale != 1.0:
                    df[col] = df[col] * scale
                    print(f"  Applied scaling factor {scale} to {col}")
    
    # Check which X-axes are available in the data
    available_x_axes = [x for x in x_axes if x in df.columns]
    if not available_x_axes:
        print(f"  Warning: None of the selected X-axes {x_axes} found in data. Skipping file.")
        return None
    
    if len(available_x_axes) < len(x_axes):
        missing = [x for x in x_axes if x not in df.columns]
        print(f"  Warning: {', '.join(missing)} not found in data. Will create graphs for available X-axes only.")
    
    print(f"  Creating graphs for {len(selected_columns)} columns with {len(available_x_axes)} X-axis type(s)...")
    
    # Create output filename
    output_path = csv_path.replace('.csv', '_analysis.xlsx')
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Add raw data sheet
    data_sheet = wb.create_sheet("Raw Data")
    for r in dataframe_to_rows(df, index=False, header=True):
        data_sheet.append(r)
    
    # Create graphs for selected columns with each X-axis
    chart_count = 0
    for x_axis in available_x_axes:
        x_col_idx = df.columns.get_loc(x_axis) + 1
        
        # Determine suffix for sheet name based on X-axis
        x_suffix = "_Time" if x_axis == "Time (sec)" else "_RPM"
        
        for col in selected_columns:
            if col in df.columns and col not in x_axes:  # Don't graph X-axis vs itself
                try:
                    col_idx = df.columns.get_loc(col) + 1
                    safe_name = col.replace('/', '_').replace('(', '').replace(')', '').replace('#', '').replace(' ', '_').replace(':', '_')
                    
                    # Add X-axis suffix to differentiate between Time and RPM graphs
                    if len(available_x_axes) > 1:
                        sheet_name = f"{safe_name}{x_suffix}"[:31]
                    else:
                        sheet_name = safe_name[:31]
                    
                    # Ensure unique sheet names
                    base_name = sheet_name
                    counter = 1
                    while sheet_name in wb.sheetnames:
                        sheet_name = f"{base_name[:28]}_{counter}"
                        counter += 1
                    
                    add_chart_to_sheet(wb, sheet_name, data_sheet, x_col_idx, col_idx, col, 
                                     df, x_axis, col)
                    chart_count += 1
                except Exception as e:
                    print(f"  Warning: Could not create chart for {col} vs {x_axis}: {e}")
    
    # Save workbook
    wb.save(output_path)
    print(f"  ✓ Saved: {output_path}")
    print(f"  ✓ Created {chart_count} total chart sheets\n")
    
    # Return dataframe for combined workbook creation
    return df

def select_files_dialog(files, title_text="Select which files to include:"):
    """
    Create a dialog for user to select which files to include
    
    Args:
        files: List of file paths
        title_text: Title text for the dialog
    
    Returns:
        List of selected file paths
    """
    selected_files = []
    
    def on_submit():
        nonlocal selected_files
        selected_files = [file for file, var in checkboxes.items() if var.get()]
        root.quit()
        root.destroy()
    
    def select_all():
        for var in checkboxes.values():
            var.set(True)
    
    def deselect_all():
        for var in checkboxes.values():
            var.set(False)
    
    # Create main window
    root = tk.Tk()
    root.title("Select Files")
    root.geometry("700x600")
    
    # Title label
    title_label = tk.Label(root, text=title_text, 
                           font=("Arial", 12, "bold"))
    title_label.pack(pady=10)
    
    # Button frame
    button_frame = tk.Frame(root)
    button_frame.pack(pady=5)
    
    select_all_btn = tk.Button(button_frame, text="Select All", command=select_all)
    select_all_btn.pack(side=tk.LEFT, padx=5)
    
    deselect_all_btn = tk.Button(button_frame, text="Deselect All", command=deselect_all)
    deselect_all_btn.pack(side=tk.LEFT, padx=5)
    
    # Create scrollable frame for checkboxes
    canvas = tk.Canvas(root)
    scrollbar = ttk.Scrollbar(root, orient="vertical", command=canvas.yview)
    scrollable_frame = ttk.Frame(canvas)
    
    scrollable_frame.bind(
        "<Configure>",
        lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
    )
    
    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)
    
    # Create checkboxes
    checkboxes = {}
    for file in files:
        var = tk.BooleanVar(value=True)  # Default to selected
        checkboxes[file] = var
        cb = ttk.Checkbutton(scrollable_frame, text=file.name, variable=var)
        cb.pack(anchor='w', padx=20, pady=2)
    
    canvas.pack(side="left", fill="both", expand=True, padx=(20, 0))
    scrollbar.pack(side="right", fill="y")
    
    # Submit button
    submit_btn = tk.Button(root, text="Continue", command=on_submit, 
                          bg="green", fg="white", font=("Arial", 10, "bold"))
    submit_btn.pack(pady=20)
    
    root.mainloop()
    
    return selected_files

def ask_yes_no(question):
    """Ask a yes/no question via GUI"""
    answer = False
    
    def on_yes():
        nonlocal answer
        answer = True
        root.quit()
        root.destroy()
    
    def on_no():
        nonlocal answer
        answer = False
        root.quit()
        root.destroy()
    
    root = tk.Tk()
    root.title("Question")
    root.geometry("400x150")
    
    label = tk.Label(root, text=question, font=("Arial", 11))
    label.pack(pady=30)
    
    button_frame = tk.Frame(root)
    button_frame.pack(pady=10)
    
    yes_btn = tk.Button(button_frame, text="Yes", command=on_yes, 
                       bg="green", fg="white", font=("Arial", 10, "bold"), width=10)
    yes_btn.pack(side=tk.LEFT, padx=10)
    
    no_btn = tk.Button(button_frame, text="No", command=on_no, 
                      bg="red", fg="white", font=("Arial", 10, "bold"), width=10)
    no_btn.pack(side=tk.LEFT, padx=10)
    
    root.mainloop()
    
    return answer

def select_multiline_columns_dialog(columns):
    """
    Create a dialog for selecting columns to combine on one chart
    
    Returns:
        List of column lists (each sub-list represents one multi-line chart)
    """
    combined_groups = []
    current_group = []
    
    def on_add_to_group():
        nonlocal current_group
        for col, var in checkboxes.items():
            if var.get():
                current_group.append(col)
                var.set(False)
        update_group_display()
    
    def on_finish_group():
        nonlocal current_group, combined_groups
        if current_group:
            combined_groups.append(current_group[:])
            current_group = []
            update_group_display()
    
    def on_done():
        nonlocal current_group, combined_groups
        if current_group:
            combined_groups.append(current_group[:])
        root.quit()
        root.destroy()
    
    def update_group_display():
        group_text.delete(1.0, tk.END)
        if current_group:
            group_text.insert(tk.END, "Current group:\n" + "\n".join(current_group) + "\n\n")
        if combined_groups:
            for i, group in enumerate(combined_groups):
                group_text.insert(tk.END, f"Chart {i+1}:\n" + "\n".join(group) + "\n\n")
    
    root = tk.Tk()
    root.title("Create Multi-Line Charts")
    root.geometry("800x700")
    
    # Instructions
    inst_label = tk.Label(root, text="Select columns to combine on one chart, then click 'Finish Chart'.\nRepeat for each multi-line chart you want.",
                         font=("Arial", 10))
    inst_label.pack(pady=10)
    
    # Main frame with two columns
    main_frame = tk.Frame(root)
    main_frame.pack(fill="both", expand=True, padx=10, pady=5)
    
    # Left side - column selection
    left_frame = tk.Frame(main_frame)
    left_frame.pack(side=tk.LEFT, fill="both", expand=True, padx=5)
    
    tk.Label(left_frame, text="Available Columns:", font=("Arial", 10, "bold")).pack()
    
    canvas = tk.Canvas(left_frame)
    scrollbar = ttk.Scrollbar(left_frame, orient="vertical", command=canvas.yview)
    scrollable_frame = ttk.Frame(canvas)
    
    scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)
    
    checkboxes = {}
    for col in columns:
        if col != 'Time (sec)':
            var = tk.BooleanVar(value=False)
            checkboxes[col] = var
            cb = ttk.Checkbutton(scrollable_frame, text=col, variable=var)
            cb.pack(anchor='w', padx=10, pady=2)
    
    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")
    
    # Buttons
    btn_frame = tk.Frame(left_frame)
    btn_frame.pack(pady=10)
    
    add_btn = tk.Button(btn_frame, text="Add Selected to Chart", command=on_add_to_group,
                       bg="blue", fg="white", font=("Arial", 9, "bold"))
    add_btn.pack(side=tk.LEFT, padx=5)
    
    finish_btn = tk.Button(btn_frame, text="Finish Chart", command=on_finish_group,
                          bg="orange", fg="white", font=("Arial", 9, "bold"))
    finish_btn.pack(side=tk.LEFT, padx=5)
    
    # Right side - groups display
    right_frame = tk.Frame(main_frame)
    right_frame.pack(side=tk.RIGHT, fill="both", expand=True, padx=5)
    
    tk.Label(right_frame, text="Charts to Create:", font=("Arial", 10, "bold")).pack()
    
    group_text = tk.Text(right_frame, width=40, height=30, font=("Arial", 9))
    group_text.pack(fill="both", expand=True)
    
    # Done button
    done_btn = tk.Button(root, text="Done - Create These Charts", command=on_done,
                        bg="green", fg="white", font=("Arial", 10, "bold"))
    done_btn.pack(pady=10)
    
    root.mainloop()
    
    return combined_groups

def add_multiline_chart_to_sheet(wb, sheet_name, data_sheet, x_col_name, columns, df, title="Multi-Line Chart"):
    """
    Add a multi-line color-coded chart to a sheet with auto-normalization
    
    Args:
        wb: Workbook object
        sheet_name: Name for the chart sheet
        data_sheet: Sheet containing the data
        x_col_name: Name of x-axis column
        columns: List of column names to include as separate lines
        df: DataFrame containing the data
        title: Chart title
    """
    # Define color palette (hex colors)
    colors = ["FF0000", "0000FF", "00FF00", "FFA500", "800080", "FF00FF", 
              "00FFFF", "FFD700", "8B4513", "000080", "008000", "FF1493"]
    
    # Create new sheet for chart
    chart_sheet = wb.create_sheet(sheet_name)
    
    # Copy X-axis data
    chart_sheet.cell(1, 1, x_col_name)
    for idx, x_val in enumerate(df[x_col_name], start=2):
        chart_sheet.cell(idx, 1, x_val)
    
    # Normalize and copy each column's data
    for col_idx, col_name in enumerate(columns, start=2):
        # Get original data
        col_data = df[col_name]
        
        # Normalize to 0-100 scale for visual comparison
        min_val = col_data.min()
        max_val = col_data.max()
        
        if max_val - min_val > 0:
            # Normalize to 0-100 range
            normalized_data = 100 * (col_data - min_val) / (max_val - min_val)
        else:
            normalized_data = col_data
        
        # Add column header with original range info
        chart_sheet.cell(1, col_idx, f"{col_name} [{min_val:.1f}-{max_val:.1f}]")
        
        # Copy normalized data
        for idx, data_val in enumerate(normalized_data, start=2):
            chart_sheet.cell(idx, col_idx, data_val)
    
    # Create chart
    chart = LineChart()
    chart.title = title + " (Normalized)"
    chart.style = None
    chart.x_axis.title = x_col_name
    chart.y_axis.title = "Normalized Values (0-100)"
    chart.width = 25
    chart.height = 15
    
    # Configure X-axis
    chart.x_axis.tickLblSkip = max(1, len(df) // 6)
    chart.x_axis.number_format = '0'
    chart.x_axis.majorTickMark = "in"
    chart.x_axis.tickLblPos = "low"
    
    # Add gridlines
    chart.y_axis.majorGridlines = ChartLines()
    
    # Ensure axes visible
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    
    # Add data series for each column
    x_values = Reference(chart_sheet, min_col=1, min_row=2, max_row=len(df)+1)
    
    for col_idx, col_name in enumerate(columns):
        y_values = Reference(chart_sheet, min_col=col_idx+2, min_row=1, max_row=len(df)+1)
        chart.add_data(y_values, titles_from_data=True)
        
        # Color code each line
        if chart.series:
            series_idx = len(chart.series) - 1
            color = colors[series_idx % len(colors)]
            chart.series[series_idx].graphicalProperties.line.solidFill = color
            chart.series[series_idx].smooth = True
    
    chart.set_categories(x_values)
    
    # Add chart to sheet
    chart_sheet.add_chart(chart, "D2")

def main():
    """
    Main function to process all CSV files in the Data directory
    """
    print("=" * 60)
    print("FSAE Data Post-Processing Script")
    print("=" * 60)
    print()
    
    # Get the script directory
    script_dir = Path(__file__).parent
    
    # Find all CSV files in the directory
    csv_files = list(script_dir.glob("*.csv"))
    
    if not csv_files:
        print("No CSV files found in the Data directory!")
        return
    
    print(f"Found {len(csv_files)} CSV file(s) to process")
    
    # Read all CSV files to get complete column list
    print("\nScanning all CSV files for column names...")
    all_columns_set = set()
    
    for csv_file in csv_files:
        with open(csv_file, 'r', encoding='latin-1', errors='ignore') as f:
            lines = f.readlines()
        
        header_line = 0
        for i, line in enumerate(lines):
            if line.startswith('Time'):
                header_line = i
                break
        
        df_temp = pd.read_csv(csv_file, encoding='latin-1', skiprows=header_line, 
                             on_bad_lines='skip', nrows=1)
        df_temp.columns = df_temp.columns.str.strip()
        
        # Add valid columns to set
        for col in df_temp.columns:
            if all(ord(c) < 128 for c in str(col)) and col.strip():
                all_columns_set.add(col)
    
    # Convert to sorted list
    valid_columns = sorted(list(all_columns_set))
    print(f"Found {len(valid_columns)} unique columns across all files")
    
    # Show column selection dialog for individual analysis files
    print("\nStep 1: Select columns for individual analysis files")
    print("Opening column selection dialog...")
    selected_columns = select_columns_dialog(valid_columns, 
                                             "Select columns for individual analysis files:")
    
    if not selected_columns:
        print("\nNo columns selected. Exiting.")
        return
    
    print(f"\n✓ Selected {len(selected_columns)} columns for individual analysis")
    
    # Get X-axis selection for individual analysis
    print("\nOpening X-axis selection dialog...")
    x_axis_individual = select_x_axis_dialog()
    print(f"✓ Selected X-axis: {x_axis_individual}")
    
    # Get scaling factors for selected columns
    print("\nOpening scaling factors dialog...")
    scaling_factors = get_scaling_factors_dialog(selected_columns)
    print(f"✓ Scaling factors set")
    
    print("\nProcessing all CSV files...\n")
    
    # Process each CSV file with the same column selection, scaling factors, and X-axis
    all_dataframes = []
    for csv_file in csv_files:
        try:
            df = process_csv_file(str(csv_file), selected_columns, scaling_factors, x_axis_individual)
            if df is not None:
                all_dataframes.append((csv_file.stem, df))
        except Exception as e:
            print(f"ERROR processing {csv_file}: {e}\n")
    
    # Ask if user wants to create multi-line charts for individual files
    print("\nStep 1.5: Multi-line charts for individual files")
    create_multiline_individual = ask_yes_no("Create multi-line color-coded charts for individual files?")
    
    if create_multiline_individual:
        # Let user select which files to create multi-line charts for
        print("Opening file selection dialog...")
        analysis_files = list(script_dir.glob("*_analysis.xlsx"))
        
        if analysis_files:
            selected_multiline_files = select_files_dialog(analysis_files,
                                                          "Select files to add multi-line charts to:")
            
            if selected_multiline_files:
                print("Opening multi-line chart selection dialog...")
                multiline_groups = select_multiline_columns_dialog(selected_columns)
                
                if multiline_groups:
                    print(f"\nAdding multi-line charts to {len(selected_multiline_files)} file(s)...")
                    
                    for excel_file in selected_multiline_files:
                        try:
                            # Load the workbook and dataframe
                            from openpyxl import load_workbook
                            wb = load_workbook(excel_file)
                            df = pd.read_excel(excel_file, sheet_name="Raw Data")
                            
                            print(f"  Adding to: {excel_file.name}")
                            chart_count = 0
                            
                            for x_axis in x_axis_individual:
                                if x_axis not in df.columns:
                                    continue
                                
                                x_suffix = "_Time" if x_axis == "Time (sec)" else "_RPM"
                                
                                for group_columns in multiline_groups:
                                    valid_group_cols = [col for col in group_columns 
                                                       if col in df.columns and col not in x_axis_individual]
                                    
                                    if valid_group_cols:
                                        chart_name = "_".join([c[:5] for c in valid_group_cols[:3]])
                                        if len(x_axis_individual) > 1:
                                            safe_name = f"Multi_{chart_name}{x_suffix}"[:31]
                                        else:
                                            safe_name = f"Multi_{chart_name}"[:31]
                                        
                                        base_name = safe_name
                                        counter = 1
                                        while safe_name in wb.sheetnames:
                                            safe_name = f"{base_name[:28]}_{counter}"
                                            counter += 1
                                        
                                        title = f"Combined: {', '.join(valid_group_cols[:3])}"
                                        if len(valid_group_cols) > 3:
                                            title += f" (+{len(valid_group_cols)-3} more)"
                                        
                                        add_multiline_chart_to_sheet(wb, safe_name, None, x_axis,
                                                                    valid_group_cols, df, title)
                                        chart_count += 1
                            
                            wb.save(excel_file)
                            print(f"    ✓ Added {chart_count} multi-line chart(s)")
                            
                        except Exception as e:
                            print(f"  Error adding charts to {excel_file.name}: {e}")
    
    # Ask if user wants combined workbook
    print("\nStep 2: Combined workbook")
    create_combined = ask_yes_no("Create a combined workbook with selected columns?")
    
    if create_combined:
        # Find all Excel files in the directory
        excel_files = list(script_dir.glob("*_analysis.xlsx"))
        
        if not excel_files:
            print("No Excel analysis files found in the directory!")
        else:
            print(f"\nFound {len(excel_files)} Excel analysis file(s)")
            
            # Let user select which files to include
            print("Opening file selection dialog...")
            selected_excel_files = select_files_dialog(excel_files, 
                                                       "Select which Excel files to include in combined workbook:")
            
            if not selected_excel_files:
                print("No files selected for combined workbook.")
            else:
                print(f"\n✓ Selected {len(selected_excel_files)} files")
                
                # Load data from selected Excel files
                excel_dataframes = []
                for excel_file in selected_excel_files:
                    try:
                        # Read the "Raw Data" sheet from each Excel file
                        df = pd.read_excel(excel_file, sheet_name="Raw Data")
                        excel_dataframes.append((excel_file.stem.replace('_analysis', ''), df))
                        print(f"  Loaded: {excel_file.name}")
                    except Exception as e:
                        print(f"  Warning: Could not load {excel_file.name}: {e}")
                
                if not excel_dataframes:
                    print("\nNo data could be loaded from selected files.")
                else:
                    # Gather all columns from selected files
                    excel_columns_set = set()
                    for _, df in excel_dataframes:
                        for col in df.columns:
                            if col != 'Time (sec)':
                                excel_columns_set.add(col)
                    
                    excel_columns = sorted(list(excel_columns_set))
                    
                    # Show column selection for combined workbook
                    print("\nOpening column selection for combined workbook...")
                    combined_columns = select_columns_dialog(excel_columns,
                                                             "Select columns for combined workbook:")
                    
                    if not combined_columns:
                        print("No columns selected for combined workbook.")
                    else:
                        # Get X-axis selection for combined workbook
                        print("\nOpening X-axis selection dialog for combined workbook...")
                        x_axis_combined = select_x_axis_dialog()
                        print(f"✓ Selected X-axis: {x_axis_combined}")
                        
                        # Get scaling factors for combined workbook
                        print("\nOpening scaling factors dialog for combined workbook...")
                        combined_scaling_factors = get_scaling_factors_dialog(combined_columns)
                        print(f"✓ Scaling factors set for combined workbook")
                        
                        # Apply scaling factors to dataframes
                        for file_name, df in excel_dataframes:
                            for col in combined_columns:
                                if col in df.columns and col in combined_scaling_factors:
                                    scale = combined_scaling_factors[col]
                                    if scale != 1.0:
                                        df[col] = df[col] * scale
                                        print(f"  Applied scaling factor {scale} to {col} in {file_name}")
                        
                        # Check which X-axes are available in all dataframes
                        available_x_axes_combined = []
                        for x_axis in x_axis_combined:
                            missing_files = [name for name, df in excel_dataframes if x_axis not in df.columns]
                            if missing_files:
                                print(f"\n  Warning: {x_axis} not found in: {', '.join(missing_files)}")
                            else:
                                available_x_axes_combined.append(x_axis)
                        
                        if not available_x_axes_combined:
                            print("\n  No valid X-axes found in all selected files.")
                        else:
                            print(f"  Will create graphs for X-axes: {', '.join(available_x_axes_combined)}")
                            
                            # Get name for combined workbook
                            combined_name = get_combined_workbook_name()
                            
                            print(f"\n{'=' * 60}")
                            print(f"Creating combined workbook: {combined_name}.xlsx")
                            print(f"{'=' * 60}")
                            
                            combined_wb = Workbook()
                            combined_wb.remove(combined_wb.active)
                            
                            # For each X-axis, file, and selected column, create a separate sheet
                            chart_count = 0
                            for x_axis in available_x_axes_combined:
                                # Determine suffix for sheet name based on X-axis
                                x_suffix = "_Time" if x_axis == "Time (sec)" else "_RPM"
                                
                                for file_name, df in excel_dataframes:
                                    if x_axis not in df.columns:
                                        continue
                                    
                                    for col in combined_columns:
                                        if col not in x_axis_combined and col in df.columns:
                                            print(f"  Adding sheet: {file_name} - {col} vs {x_axis}")
                                            
                                            # Create safe sheet name
                                            base_safe_name = f"{file_name}_{col}".replace('/', '_').replace('(', '').replace(')', '').replace('#', '').replace(' ', '_').replace(':', '_')
                                            
                                            # Add X-axis suffix if multiple X-axes selected
                                            if len(available_x_axes_combined) > 1:
                                                safe_name = f"{base_safe_name}{x_suffix}"[:31]
                                            else:
                                                safe_name = base_safe_name[:31]
                                            
                                            # Ensure unique sheet names
                                            sheet_name = safe_name
                                            counter = 1
                                            while sheet_name in combined_wb.sheetnames:
                                                sheet_name = f"{safe_name[:28]}_{counter}"
                                                counter += 1
                                            
                                            sheet = combined_wb.create_sheet(sheet_name)
                                            
                                            # Copy data to sheet (scaled data)
                                            sheet.cell(1, 1, x_axis)
                                            sheet.cell(1, 2, col)
                                            
                                            for idx, (x_val, data_val) in enumerate(zip(df[x_axis], df[col]), start=2):
                                                sheet.cell(idx, 1, x_val)
                                                sheet.cell(idx, 2, data_val)
                                            
                                            # Create chart
                                            chart = LineChart()
                                            chart.title = f"{file_name} - {col}"
                                            chart.style = None
                                            chart.x_axis.title = x_axis
                                            chart.y_axis.title = col
                                            chart.width = 25   # Very large chart to give more room for axis titles
                                            chart.height = 15  # Very tall chart to give more room for axis titles
                                            
                                            # Show only ~5-6 labels across the entire x-axis (similar to y-axis)
                                            chart.x_axis.tickLblSkip = max(1, len(df) // 6)  # Show roughly 6 labels total
                                            chart.x_axis.number_format = '0'  # Format as whole numbers (no decimals)
                                            chart.x_axis.majorTickMark = "in"  # Add inside tick marks
                                            chart.x_axis.tickLblPos = "low"  # Position labels below axis
                                            
                                            # Add only horizontal gridlines (Y-axis gridlines only)
                                            chart.y_axis.majorGridlines = ChartLines()
                                            
                                            # Ensure axes are always visible
                                            chart.x_axis.delete = False
                                            chart.y_axis.delete = False
                                            
                                            chart.legend = None
                                            
                                            # Add data
                                            y_values = Reference(sheet, min_col=2, min_row=1, max_row=len(df)+1)
                                            x_values = Reference(sheet, min_col=1, min_row=2, max_row=len(df)+1)
                                            chart.add_data(y_values, titles_from_data=True)
                                            chart.set_categories(x_values)
                                            
                                            # Simple black line
                                            if chart.series:
                                                chart.series[0].graphicalProperties.line.solidFill = "000000"
                                                chart.series[0].smooth = True
                                            
                                            # Position chart
                                            sheet.add_chart(chart, "D2")
                                            
                                            chart_count += 1
                        
                            # Ask if user wants multi-line charts in combined workbook
                            print("\n  Do you want to create multi-line color-coded charts in combined workbook?")
                            create_multiline_combined = ask_yes_no("Create multi-line charts for combined workbook?")
                            
                            if create_multiline_combined:
                                # Let user select which files to use for multi-line charts
                                print("  Select which file(s) to use for multi-line charts...")
                                file_dict = {f"{name}": (name, df) for name, df in excel_dataframes}
                                selected_multiline_files = select_files_dialog(
                                    [Path(name + "_analysis.xlsx") for name in file_dict.keys()],
                                    "Select files to use for multi-line charts:")
                                
                                if selected_multiline_files:
                                    # Get the dataframes for selected files
                                    selected_df_data = [(file_dict[f.stem.replace('_analysis', '')][0], 
                                                        file_dict[f.stem.replace('_analysis', '')][1]) 
                                                       for f in selected_multiline_files 
                                                       if f.stem.replace('_analysis', '') in file_dict]
                                    
                                    print("  Opening multi-line chart selection dialog...")
                                    multiline_groups_combined = select_multiline_columns_dialog(combined_columns)
                                    
                                    if multiline_groups_combined:
                                        print(f"  Creating {len(multiline_groups_combined)} multi-line chart(s)...")
                                        
                                        for x_axis in available_x_axes_combined:
                                            x_suffix = "_Time" if x_axis == "Time (sec)" else "_RPM"
                                            
                                            for file_name, df in selected_df_data:
                                                if x_axis not in df.columns:
                                                    continue
                                                
                                                for group_idx, group_columns in enumerate(multiline_groups_combined):
                                                    # Filter columns that exist in this dataframe
                                                    valid_group_cols = [col for col in group_columns 
                                                                       if col in df.columns and col not in x_axis_combined]
                                                    
                                                    if valid_group_cols:
                                                        # Create chart name
                                                        chart_name = "_".join([c[:5] for c in valid_group_cols[:3]])
                                                        if len(available_x_axes_combined) > 1:
                                                            safe_name = f"{file_name}_Multi_{chart_name}{x_suffix}"[:31]
                                                        else:
                                                            safe_name = f"{file_name}_Multi_{chart_name}"[:31]
                                                        
                                                        # Ensure unique sheet names
                                                        base_name = safe_name
                                                        counter = 1
                                                        while safe_name in combined_wb.sheetnames:
                                                            safe_name = f"{base_name[:28]}_{counter}"
                                                            counter += 1
                                                        
                                                        try:
                                                            title = f"{file_name} - Combined: {', '.join(valid_group_cols[:3])}"
                                                            if len(valid_group_cols) > 3:
                                                                title += f" (+{len(valid_group_cols)-3} more)"
                                                            
                                                            # Use the normalized version
                                                            add_multiline_chart_to_sheet(combined_wb, safe_name, None, x_axis, 
                                                                                        valid_group_cols, df, title)
                                                            chart_count += 1
                                                            print(f"    Created: {title}")
                                                        except Exception as e:
                                                            print(f"    Warning: Could not create multi-line chart: {e}")
                            
                            # Save combined workbook
                            combined_path = script_dir / f"{combined_name}.xlsx"
                            combined_wb.save(combined_path)
                            print(f"\n  ✓ Saved combined workbook: {combined_path}")
                            print(f"  ✓ Created {chart_count} total sheets\n")
    
    print("=" * 60)
    print("Processing complete!")
    print("=" * 60)

if __name__ == "__main__":
    main()
